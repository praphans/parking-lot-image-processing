import os
import json
import cv2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from ultralytics import YOLO

# Define parameters for adjustment
brightness_values = [10, 20, 30, 40, 50, 60]
contrast_values = [5, 10, 15, 20, 25, 30]
noise_values = [0.5, 1, 1.5, 2, 2.5, 3]

# Load the YOLO model
model = YOLO('yolov8s.pt')
# model = YOLO('runs/detect/train/weights/best.pt')

# Function to adjust brightness and contrast
def adjust_brightness_contrast(image, brightness=0, contrast=0):
    # Brightness
    if brightness != 0:
        if brightness > 0:
            shadow = brightness
            highlight = 255
        else:
            shadow = 0
            highlight = 255 + brightness
        alpha_b = (highlight - shadow) / 255
        gamma_b = shadow
        buf = cv2.addWeighted(image, alpha_b, image, 0, gamma_b)
    else:
        buf = image.copy()

    # Contrast
    if contrast != 0:
        f = 131 * (contrast + 127) / (127 * (131 - contrast))
        alpha_c = f
        gamma_c = 127 * (1 - f)
        buf = cv2.addWeighted(buf, alpha_c, buf, 0, gamma_c)

    return buf

# Function to reduce noise
def reduce_noise(image, noise_value):
    return cv2.fastNlMeansDenoisingColored(image, None, noise_value, noise_value, 7, 21)

# Read "coco.txt" file and split the data into a list of classes
with open("coco.txt", "r") as my_file:
    data = my_file.read()
    class_list = data.split("\n")

# Function to extract bounding boxes around objects
def get_boxes(results, class_id):
    boxes = []
    for result in results:
        for det in result.boxes:
            if int(det.cls) == class_id:
                # ย้ายเทนเซอร์ไปยัง CPU ก่อนแปลงเป็น NumPy array
                boxes.append(det.xyxy.cpu().numpy())
    return boxes

# Get the list of images in the folder
image_folder = 'images'
image_files = [f for f in os.listdir(image_folder) if f.endswith('.jpg')]

# Array to hold the results
results_list = []

for image_file in image_files:
    image_path = os.path.join(image_folder, image_file)
    frame = cv2.imread(image_path)  # Normal parking lot image
    
    # Process results for each combination of brightness, contrast, and noise
    for brightness_value in brightness_values:
        for contrast_value in contrast_values:
            for noise_value in noise_values:
                # Adjust the brightness and contrast of the image
                frame_adjusted = adjust_brightness_contrast(frame, brightness_value, contrast_value)

                # Reduce noise in the image
                frame_adjusted = reduce_noise(frame_adjusted, noise_value)

                # Detect objects in the image
                results = model(frame_adjusted)

                # Extract the class ID for cars
                car_class_id = class_list.index('car')

                # Get bounding boxes for cars
                car_boxes = get_boxes(results, car_class_id)

                # Count the number of cars
                num_cars = len(car_boxes)

                # Append the result to the list
                results_list.append({
                    'image': image_file,
                    'brightness_value': brightness_value,
                    'contrast_value': contrast_value,
                    'noise_value': noise_value,
                    'cars': num_cars
                })

# Save the results to a JSON file
with open('brightness_contrast_noise.json', 'w') as outfile:
    json.dump(results_list, outfile, indent=4)

# Convert the results list to a DataFrame
df = pd.DataFrame(results_list)

# Save the results to an Excel file
excel_path = 'brightness_contrast_noise.xlsx'
df.to_excel(excel_path, index=False)

# Load the Excel file
wb = load_workbook(excel_path)
ws = wb.active

# Merge cells for each unique image
current_image = None
start_row = 2  # Starting from the first data row

for index, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
    image = row[0]
    if image != current_image:
        if current_image is not None:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=index-1, end_column=1)
            for merged_row in range(start_row, index):
                cell = ws.cell(row=merged_row, column=1)
                cell.alignment = Alignment(vertical='center')
        current_image = image
        start_row = index

# Merge cells for the last image
ws.merge_cells(start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1)
for merged_row in range(start_row, ws.max_row + 1):
    cell = ws.cell(row=merged_row, column=1)
    cell.alignment = Alignment(vertical='center')

# Save the updated Excel file
wb.save(excel_path)

print("Results saved to 'brightness_contrast_noise.json' and 'brightness_contrast_noise.xlsx'")
